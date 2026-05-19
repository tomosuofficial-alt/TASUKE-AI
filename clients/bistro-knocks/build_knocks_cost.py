from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

OUT = "/Users/takashiouchi/Developer/TASUKE-AI/.claude/worktrees/peaceful-tu-b43503/clients/bistro-knocks/bistro_knocks_cost_master.xlsx"

# ===== デザイン定数（Bistro Knocks 提出版）=====
FONT = "Hiragino Sans"  # 日本語対応モダン。Google Sheetsでも代替表示可能
FONT_FALLBACK = "Arial"

# パレット: チャコール×ウォームゴールド（ビストロ調）
COLOR_HEADER_BG = "2C3E50"       # チャコール
COLOR_HEADER_FG = "FFFFFF"       # 白
COLOR_SECTION_BG = "EFEBE0"      # ウォームベージュ
COLOR_INPUT_BG = "FFF8E1"        # アイボリー（淡い）
COLOR_COMPUTED_BG = "FFFFFF"     # 白
COLOR_TOTAL_BG = "FAE6B0"        # 淡ゴールド（合計強調）
COLOR_BORDER = "D5D5D5"          # ライトグレー
COLOR_ACCENT = "B8860B"          # ダークゴールド（タイトル下線等）

INPUT_FILL = PatternFill("solid", start_color=COLOR_INPUT_BG)
HEADER_FILL = PatternFill("solid", start_color=COLOR_HEADER_BG)
HEADER_FONT = Font(name=FONT, bold=True, color=COLOR_HEADER_FG, size=11)
SECTION_FILL = PatternFill("solid", start_color=COLOR_SECTION_BG)
TOTAL_FILL = PatternFill("solid", start_color=COLOR_TOTAL_BG)
BASE_FONT = Font(name=FONT, size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=20, color=COLOR_HEADER_BG)
SUBTITLE_FONT = Font(name=FONT, bold=True, size=12, color=COLOR_ACCENT)
THIN = Side(style="thin", color=COLOR_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK_BOTTOM = Border(bottom=Side(style="medium", color=COLOR_ACCENT))

# 原価率の条件付き書式ルール
RULE_GREEN  = CellIsRule(operator='lessThan', formula=['0.25'], fill=PatternFill("solid", start_color="C6EFCE"), font=Font(name=FONT, color="006100", bold=True))
RULE_NORMAL = CellIsRule(operator='between',  formula=['0.25','0.35'], fill=PatternFill("solid", start_color="FFFFFF"))
RULE_WARN   = CellIsRule(operator='between',  formula=['0.35','0.45'], fill=PatternFill("solid", start_color="FFEB9C"), font=Font(name=FONT, color="9C5700", bold=True))
RULE_DANGER = CellIsRule(operator='greaterThan', formula=['0.45'], fill=PatternFill("solid", start_color="FFC7CE"), font=Font(name=FONT, color="9C0006", bold=True))

wb = Workbook()

# -------------- 食材マスタ --------------
ws_ing = wb.active
ws_ing.title = "食材マスタ"
ingredients = [
    ("150gバーグ", "個", "黒弁当・新弁当で使用"),
    ("120gバーグ", "個", "新弁当で主に使用"),
    ("ハラミ", "g", "最も使用頻度が高い"),
    ("ロービー", "g", "複数メニューで使用"),
    ("ミスジ", "g", "高級メニューで使用"),
    ("イベリコ", "g", "プレミアムメニュー"),
    ("厳選和牛", "g", "最高級メニュー"),
    ("豚ヒレ", "g", "新弁当で使用"),
    ("豚ロース", "g", "新弁当で使用"),
    ("とりむね", "g", "新弁当で使用"),
    ("鯛", "カット", "黒弁当・新弁当で使用"),
    ("サーモン", "カット", "黒弁当・新弁当で使用"),
    ("カジキ", "g", "新弁当で使用"),
    ("赤魚", "g", "新弁当で使用"),
    ("大豆バーグ", "個", "低糖質弁当"),
    ("豆腐", "丁", "低糖質弁当"),
]
# 食材マスタは「途中計算」を見える形に: 仕入単価(税込) → 仕入単位量 → 税込/単位 → 税抜/単位
ws_ing.append([
    "食材名", "使用単位",
    "仕入単価（税込・円）", "仕入単位量",
    "税込単位単価（円/使用単位）", "税抜単位単価（円/使用単位）",
    "備考",
])
for r in ingredients:
    ws_ing.append([r[0], r[1], None, None, None, None, r[2]])

# 業者単価入力
# 豚ヒレ・豚ロース（2026-05-19 入手分・税込/100g・掃除前）
ws_ing.cell(row=9, column=3, value=117)   # 豚ヒレ
ws_ing.cell(row=9, column=4, value=100)
ws_ing.cell(row=9, column=7, value="新弁当で使用 / 117円/100g税込・掃除前（2026-05-19入手）")
ws_ing.cell(row=10, column=3, value=98)   # 豚ロース
ws_ing.cell(row=10, column=4, value=100)
ws_ing.cell(row=10, column=7, value="新弁当で使用 / 98円/100g税込・掃除前（2026-05-19入手）")
# 150gバーグ（赤城牛ハンバーグ）2026-05-01 納品（レーベン）税抜420円/枚 → 税込
ws_ing.cell(row=2, column=3, value=453.6)   # 420×1.08
ws_ing.cell(row=2, column=4, value=1)
ws_ing.cell(row=2, column=7, value="赤城牛ハンバーグ150g / 税抜420円/枚（レーベン 2026-05-01）")
# 120gバーグ（ハンバーグ120g）2026-05-08/12 納品（レーベン）税抜225円/枚 → 税込
ws_ing.cell(row=3, column=3, value=243)   # 225×1.08
ws_ing.cell(row=3, column=4, value=1)
ws_ing.cell(row=3, column=7, value="ハンバーグ120g / 税抜225円/枚（レーベン 2026-05-08/12）")

# E列(税込単位単価) と F列(税抜単位単価) は全行に計算式
for r in range(2, 2 + len(ingredients)):
    ws_ing.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)")
    ws_ing.cell(row=r, column=6, value=f"=IFERROR(E{r}/1.08,0)")

# header style
for c in range(1, 8):
    cell = ws_ing.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
# input cells (C, D) - yellow / calc cells (E, F) - 通常
for r in range(2, 2 + len(ingredients)):
    ws_ing.cell(row=r, column=3).fill = INPUT_FILL
    ws_ing.cell(row=r, column=3).number_format = '#,##0;(#,##0);-'
    ws_ing.cell(row=r, column=4).fill = INPUT_FILL
    ws_ing.cell(row=r, column=4).number_format = '#,##0;(#,##0);-'
    ws_ing.cell(row=r, column=5).number_format = '#,##0.0000;(#,##0.0000);-'
    ws_ing.cell(row=r, column=6).number_format = '#,##0.0000;(#,##0.0000);-'
ws_ing.column_dimensions["A"].width = 16
ws_ing.column_dimensions["B"].width = 10
ws_ing.column_dimensions["C"].width = 18
ws_ing.column_dimensions["D"].width = 12
ws_ing.column_dimensions["E"].width = 22
ws_ing.column_dimensions["F"].width = 22
ws_ing.column_dimensions["G"].width = 38

# -------------- 飲料マスタ --------------
ws_drink = wb.create_sheet("飲料マスタ")
drinks = [
    ("none", "無し", 0, 0, "飲料なし"),
    ("pet_free", "PET茶（無料）", 0, None, "2,500円以上の弁当に自動付与"),
    ("pet", "PET茶 180円", 180, None, "有料オプション"),
    ("pack", "紙茶 100円", 100, None, "有料オプション"),
]
ws_drink.append(["飲料ID", "飲料名", "販売価格（税込・円）", "税抜原価（円）", "備考"])
for r in drinks:
    ws_drink.append(list(r))
for c in range(1, 6):
    cell = ws_drink.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
# none has 0 cost (fixed), others are input
for r in range(3, 6):  # rows 3-5: pet_free, pet, pack
    ws_drink.cell(row=r, column=4).fill = INPUT_FILL
    ws_drink.cell(row=r, column=4).number_format = '#,##0;(#,##0);-'
ws_drink.column_dimensions["A"].width = 12
ws_drink.column_dimensions["B"].width = 18
ws_drink.column_dimensions["C"].width = 16
ws_drink.column_dimensions["D"].width = 14
ws_drink.column_dimensions["E"].width = 32

# -------------- サイドマスタ --------------
ws_side = wb.create_sheet("サイドマスタ")
# 重量ベース11サイド構造
# (name, unit, qty_black, qty_new, unit_price_value_or_formula(税抜/単位), 備考)
sides = [
    # 7品 既存（フラグ管理あり、メニュー側で個別オン/オフ）
    ("ピクルス",     "g",  15,  15, "FORMULA:ピクルス",        "全メニュー / 自家製レシピ(ソース定義マスタから)"),
    ("ポテトサラダ", "g",  25,  25, None,                       "全メニュー / 自家製レシピ未登録"),
    ("卵",          "個",  1,   1, None,                       "全メニュー / 単価入力"),
    ("ご飯",        "g", 160, 160, None,                       "ほぼ全メニュー / 単価入力（低糖質はフラグで除外）"),
    ("かぼちゃ",     "g",  None, 0, None,                       "黒弁当のみ / 単価入力"),
    ("小エビ",      "個", None, 0, None,                       "黒弁当の魚系以外 / 単価入力"),
    ("有頭エビ",     "個", None, 0, None,                       "魚系メニュー / 単価入力"),
    # 4品 新規（全メニュー固定付与・フラグなし）
    ("サラダベース", "g",  35,  20, "FORMULA:サラダベース",     "全メニュー / 自家製レシピ(ソース定義マスタから)"),
    ("ミニトマト",   "個", 0.5, 0.5, "FORMULA:SUB:ミニトマト",   "全メニュー / レシピ材料マスタから直接"),
    ("蒸し野菜",     "g",  55,  35, None,                        "全メニュー / 構成未確定（後でレシピ材料/レシピ化）"),
    ("ブロッコリー", "g",   6,   6, "FORMULA:SUB:ブロッコリー(冷凍)", "全メニュー / ユニフーズ冷凍500g"),
]
ws_side.append(["サイド名", "単位", "黒弁当量", "新弁当量", "税抜単位単価（円/単位）", "備考"])
for r in sides:
    ws_side.append([r[0], r[1], r[2], r[3], None, r[5]])

for c in range(1, 7):
    cell = ws_side.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# E列に式 or 入力欄
for i, r in enumerate(sides):
    rr = i + 2
    val = r[4]
    if val is None:
        ws_side.cell(row=rr, column=5).fill = INPUT_FILL
    elif isinstance(val, str) and val.startswith("FORMULA:SUB:"):
        sub_name = val.split(":", 2)[2]
        # レシピ材料マスタ F列(税抜単位単価)からVLOOKUP（フルカラム参照で安全）
        ws_side.cell(row=rr, column=5, value=f'=IFERROR(VLOOKUP("{sub_name}",レシピ材料マスタ!$A:$F,6,FALSE),0)')
    elif isinstance(val, str) and val.startswith("FORMULA:"):
        sauce_name = val.split(":", 1)[1]
        # ソース定義マスタ E列(単位原価税抜)からVLOOKUP
        ws_side.cell(row=rr, column=5, value=f'=IFERROR(VLOOKUP("{sauce_name}",ソース定義マスタ!$A:$E,5,FALSE),0)')
    ws_side.cell(row=rr, column=5).number_format = '#,##0.0000;(#,##0.0000);-'
    ws_side.cell(row=rr, column=3).number_format = '#,##0.00;(#,##0.00);-'
    ws_side.cell(row=rr, column=4).number_format = '#,##0.00;(#,##0.00);-'

ws_side.column_dimensions["A"].width = 16
ws_side.column_dimensions["B"].width = 6
ws_side.column_dimensions["C"].width = 10
ws_side.column_dimensions["D"].width = 10
ws_side.column_dimensions["E"].width = 22
ws_side.column_dimensions["F"].width = 38

# -------------- 資材マスタ --------------
ws_mat = wb.create_sheet("資材マスタ")
materials = [
    ("容器（黒）", "サイズ別", "黒弁当のみ"),
    ("容器（新）", "サイズ別", "新弁当のみ"),
    ("蓋（黒）", "サイズ別", "黒弁当のみ"),
    ("蓋（新）", "サイズ別", "新弁当のみ"),
    ("箸", "共通", "全メニュー"),
    ("ワサガード", "共通", "全メニュー"),
    ("帯", "共通", "全メニュー"),
    ("シーザードレッシング", "共通", "全メニュー"),
    ("おしぼり", "共通", "全メニュー"),
]
ws_mat.append(["資材名", "分類", "税抜単価（円／個）", "適用"])
for r in materials:
    ws_mat.append([r[0], r[1], None, r[2]])
# シーザードレッシング（トーホー / QPシーザーサラダドレッシング 10ml×40袋/箱 税抜646円/箱 → 1袋税抜16.15円）
ws_mat.cell(row=9, column=3, value=16.15)
ws_mat.cell(row=9, column=4, value="全メニュー / QP 1袋10ml×40袋/箱 税抜646円（トーホー 2026-05-18）")
for c in range(1, 5):
    cell = ws_mat.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
for r in range(2, 2 + len(materials)):
    ws_mat.cell(row=r, column=3).fill = INPUT_FILL
    ws_mat.cell(row=r, column=3).number_format = '#,##0;(#,##0);-'
ws_mat.column_dimensions["A"].width = 22
ws_mat.column_dimensions["B"].width = 10
ws_mat.column_dimensions["C"].width = 16
ws_mat.column_dimensions["D"].width = 18

# -------------- レシピ材料マスタ（ソース・ピクルス用のサブ材料） --------------
ws_sub = wb.create_sheet("レシピ材料マスタ")
# 名前は「(単位)付き」で重複回避（例: 玉ねぎ(個) と 玉ねぎ(g) を別行）
sub_materials = [
    ("味醂", "cc", "ソース1/6で使用"),
    ("醤油", "cc", "ソース1/6で使用"),
    ("マギーブイヨン", "個", "ソース1で使用"),
    ("顆粒だし", "g", "ソース1で使用"),
    ("しめじ", "pc", "ソース1で使用（コンカッセ）"),
    ("えのき", "pc", "ソース1で使用（コンカッセ）"),
    ("エリンギ", "pc", "ソース1で使用（コンカッセ）"),
    ("水", "cc", "ソース1で使用 ※単価0想定"),
    ("粒マスタード", "g", "ソース1で使用（ソース全量×5%）"),
    ("白ワイン", "ml", "ソース2/6で使用"),
    ("クラム", "g", "ソース2で使用"),
    ("生クリーム", "ml", "ソース2で使用"),
    ("オニオンステーキソース", "cc", "ソース3で使用（市販）"),
    ("ニンニク醤油ソース", "cc", "ソース3で使用（市販）"),
    ("玉ねぎ(個)", "個", "ソース4(ピクルス)で使用"),
    ("玉ねぎ(g)", "g", "ソース6で使用"),
    ("生姜", "g", "ソース6で使用"),
    ("にんにく(株)", "株", "ソース6で使用"),
    ("日本酒", "cc", "ソース6で使用"),
    ("上白糖", "g", "ソース6で使用"),
    ("レモン", "個", "ソース6で使用"),
    ("かんたん酢", "ml", "ピクルスで使用"),
    ("白ワインビネガー", "ml", "ピクルスで使用"),
    ("塩", "g", "ピクルスで使用"),
    ("グラニュー糖", "g", "ピクルスで使用"),
    ("とうがらし", "本", "ピクルスで使用"),
    ("白粒こしょう", "粒", "ピクルスで使用"),
    ("大根", "本", "ピクルスで使用"),
    ("人参", "本", "ピクルスで使用"),
    ("パプリカ", "個", "ピクルスで使用"),
    ("ヤングコーン(缶)", "缶", "ピクルスで使用"),
    ("きゅうり", "本", "ピクルスで使用"),
    # 5. ポリネシアンソース用
    ("サラダ油", "cc", "ソース5で使用"),
    ("ケチャップ", "g", "ソース5で使用"),
    ("りんご", "個", "ソース5で使用"),
    ("にんにく(個)", "個", "ソース5で使用 ※ソース6は(株)単位で別行"),
    ("チリペッパーパウダー", "g", "ソース5で使用（適量・初期0）"),
    # 7. エバラおろしのたれ（市販ソース）
    ("エバラおろしのたれ", "cc", "ソース7（市販品）"),
    # サイド・追加レシピ材料
    ("牛モモ(加熱用)", "g", "ロービー用の原料肉（オーストラリア産）※歩留まり別途設定要"),
    ("ブロッコリー(冷凍)", "g", "サイド「ブロッコリー」用（ユニフーズ500g）"),
    ("ミニトマト", "個", "サイド「ミニトマト」用"),
    ("サニーレタス", "個", "サラダベース材料"),
    ("水菜", "束", "サラダベース材料"),
    ("白菜", "個", "サラダベース材料（ハーフで1/2）"),
    ("紫キャベツ", "個", "サラダベース材料（ハーフで1/2）"),
    ("キャベツ", "個", "蒸し野菜材料"),
    ("もやし", "g", "蒸し野菜材料"),
]
ws_sub.append([
    "材料名", "単位",
    "仕入単価（税込・円）", "仕入単位量",
    "税込単位単価（円/単位）", "税抜単位単価（円/単位）",
    "備考",
])
for r in sub_materials:
    ws_sub.append([r[0], r[1], None, None, None, None, r[2]])

# 水は単価0固定
for r in range(2, 2 + len(sub_materials)):
    if ws_sub.cell(row=r, column=1).value == "水":
        ws_sub.cell(row=r, column=3, value=0)
        ws_sub.cell(row=r, column=4, value=1)
    ws_sub.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)")
    ws_sub.cell(row=r, column=6, value=f"=IFERROR(E{r}/1.08,0)")
    ws_sub.cell(row=r, column=3).fill = INPUT_FILL
    ws_sub.cell(row=r, column=3).number_format = '#,##0;(#,##0);-'
    ws_sub.cell(row=r, column=4).fill = INPUT_FILL
    ws_sub.cell(row=r, column=4).number_format = '#,##0;(#,##0);-'
    ws_sub.cell(row=r, column=5).number_format = '#,##0.0000;(#,##0.0000);-'
    ws_sub.cell(row=r, column=6).number_format = '#,##0.0000;(#,##0.0000);-'

# 業者単価転記（トーホー/レーベン 2026-05 納品分・税抜単価を税込換算で入力）
def _ing_row(name):
    for i, t in enumerate(sub_materials):
        if t[0] == name:
            return i + 2
    return None

price_updates = [
    # name, C(税込/単位), D(単位量), 備考追記
    ("オニオンステーキソース", 700.92, 1110, "トーホー / ステーキしょうゆ オニオン&ペッパー 1110g 税抜649円"),
    ("ニンニク醤油ソース", 680.40, 1205, "トーホー / ステーキしょうゆ にんにく風味 1205g 税抜630円"),
    ("白ワインビネガー", 334.80, 1000, "トーホー / レストランビネガー ワインタイプ 1L 税抜310円"),
    ("エバラおろしのたれ", 861.84, 1000, "トーホー / エバラ厨房応援団 おろしのたれレモン醤油味 1L 税抜798円"),
    ("牛モモ(加熱用)", 2648.16, 1000, "レーベン / オーストラリア産 牛モモ(加熱用) 5.49kg 税抜2452円/kg"),
    ("ブロッコリー(冷凍)", 270.00, 500, "トーホー / ユニフーズ ブロッコリー 500g 税抜250円"),
]
for nm, c_val, d_val, note in price_updates:
    rr = _ing_row(nm)
    if rr:
        ws_sub.cell(row=rr, column=3, value=c_val)
        ws_sub.cell(row=rr, column=4, value=d_val)
        existing_note = ws_sub.cell(row=rr, column=7).value or ""
        ws_sub.cell(row=rr, column=7, value=f"{existing_note} / {note}")

for c in range(1, 8):
    cell = ws_sub.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
ws_sub.column_dimensions["A"].width = 22
ws_sub.column_dimensions["B"].width = 8
ws_sub.column_dimensions["C"].width = 18
ws_sub.column_dimensions["D"].width = 12
ws_sub.column_dimensions["E"].width = 22
ws_sub.column_dimensions["F"].width = 22
ws_sub.column_dimensions["G"].width = 30

SUB_LAST_ROW = 1 + len(sub_materials)
SUB_RANGE_NAME = f"レシピ材料マスタ!$A$2:$A${SUB_LAST_ROW}"
SUB_RANGE_PRICE = f"レシピ材料マスタ!$A$2:$F${SUB_LAST_ROW}"

# -------------- ソースレシピマスタ --------------
ws_rcp = wb.create_sheet("ソースレシピマスタ")
# 1行 = (ソース名, 材料名, 分量, 単位, 備考)
# 分量・単位はレシピ材料マスタの単位と合わせる前提（VLOOKUPで単位原価を取得）
# 基準: バッチ単位(=レシピ通り) を「batch」、1食分の場合は「per_serving」と備考に書く
recipes = [
    # 1. ハンバーグソース（バッチ）
    ("ハンバーグソース", "味醂", 1080, "cc", "アルコールを飛ばす"),
    ("ハンバーグソース", "醤油", 540, "cc", ""),
    ("ハンバーグソース", "マギーブイヨン", 3, "個", ""),
    ("ハンバーグソース", "顆粒だし", 35, "g", ""),
    ("ハンバーグソース", "しめじ", 4, "pc", "コンカッセ"),
    ("ハンバーグソース", "えのき", 2, "pc", "コンカッセ"),
    ("ハンバーグソース", "エリンギ", 2, "pc", "コンカッセ"),
    ("ハンバーグソース", "水", 1500, "cc", ""),
    # 粒マスタードはソース全量×5%。後で式で上書き
    ("ハンバーグソース", "粒マスタード", None, "g", "ソース全量×5%（自動計算）"),
    # 2. ブールブランソース（1食分）
    ("ブールブランソース", "白ワイン", 15, "ml", "1食ぶん"),
    ("ブールブランソース", "クラム", 2.5, "g", "1食ぶん"),
    ("ブールブランソース", "生クリーム", 20, "ml", "1食ぶん"),
    # 3. オニオンソース（バッチ）
    ("オニオンソース", "オニオンステーキソース", 2000, "cc", "2L=2000cc"),
    ("オニオンソース", "ニンニク醤油ソース", 300, "cc", ""),
    # 4. ピクルス（バッチ）
    ("ピクルス", "かんたん酢", 6000, "ml", "6L=6000ml"),
    ("ピクルス", "白ワインビネガー", 1000, "ml", "1L=1000ml"),
    ("ピクルス", "塩", 175, "g", ""),
    ("ピクルス", "グラニュー糖", 150, "g", ""),
    ("ピクルス", "とうがらし", 6, "本", ""),
    ("ピクルス", "白粒こしょう", 30, "粒", ""),
    ("ピクルス", "玉ねぎ(個)", 10, "個", "追加材料"),
    ("ピクルス", "大根", 3, "本", "追加材料"),
    ("ピクルス", "人参", 10, "本", "追加材料"),
    ("ピクルス", "パプリカ", 12, "個", "追加材料"),
    ("ピクルス", "ヤングコーン(缶)", 8, "缶", "追加材料"),
    ("ピクルス", "きゅうり", 12, "本", "追加材料"),
    # 5. ポリネシアンソース（バッチ）
    ("ポリネシアンソース", "サラダ油", 1080, "cc", ""),
    ("ポリネシアンソース", "ケチャップ", 540, "g", ""),
    ("ポリネシアンソース", "白ワイン", 720, "cc", ""),
    ("ポリネシアンソース", "りんご", 4, "個", ""),
    ("ポリネシアンソース", "玉ねぎ(個)", 4, "個", ""),
    ("ポリネシアンソース", "にんにく(個)", 1, "個", ""),
    ("ポリネシアンソース", "生姜", 0, "g", "適量（未確定・要実測）"),
    ("ポリネシアンソース", "醤油", 200, "cc", "ミキシング後に追加"),
    ("ポリネシアンソース", "チリペッパーパウダー", 0, "g", "適量・提供直前（未確定）"),
    # 6. ソースジャンジャンブル（バッチ）
    ("ソースジャンジャンブル", "玉ねぎ(g)", 500, "g", "ざく切り"),
    ("ソースジャンジャンブル", "生姜", 375, "g", ""),
    ("ソースジャンジャンブル", "にんにく(株)", 1, "株", "芽をとる"),
    ("ソースジャンジャンブル", "味醂", 700, "cc", ""),
    ("ソースジャンジャンブル", "白ワイン", 270, "ml", ""),
    ("ソースジャンジャンブル", "日本酒", 270, "cc", ""),
    ("ソースジャンジャンブル", "上白糖", 75, "g", ""),
    ("ソースジャンジャンブル", "レモン", 1, "個", ""),
    ("ソースジャンジャンブル", "醤油", 900, "cc", "ミキシング後に追加"),
    # 7. エバラおろしのたれ（市販品の1L換算）
    ("エバラおろしのたれ", "エバラおろしのたれ", 1000, "cc", "市販品。1Lベースで単位原価を算出"),
    # 8. 粒マスタード（単体・条件付け）
    ("粒マスタード単品", "粒マスタード", 100, "g", "そのまま提供（g単位）"),
    # サブレシピ（サイド用）
    # サラダベース: サニーレタス6個 + 水菜6束 + 白菜0.5個 + 紫キャベツ0.5個
    ("サラダベース", "サニーレタス", 6, "個", "バッチレシピ"),
    ("サラダベース", "水菜", 6, "束", "バッチレシピ"),
    ("サラダベース", "白菜", 0.5, "個", "バッチレシピ（ハーフ）"),
    ("サラダベース", "紫キャベツ", 0.5, "個", "バッチレシピ（ハーフ）"),
    # 蒸し野菜ミックス
    ("蒸し野菜", "キャベツ", 2, "個", "バッチレシピ"),
    ("蒸し野菜", "白菜", 0.5, "個", "バッチレシピ（ハーフ）"),
    ("蒸し野菜", "人参", 3, "本", "バッチレシピ（にんじん）"),
    ("蒸し野菜", "玉ねぎ(個)", 3, "個", "バッチレシピ"),
    ("蒸し野菜", "もやし", 8000, "g", "バッチレシピ（8kg）"),
    # ポテトサラダ: レシピ未受領（プレースホルダー）
]
ws_rcp.append([
    "ソース名", "材料名", "分量", "単位",
    "税抜単位単価（円/単位）", "材料費（円）",
    "出来上がり寄与量（cc/g）", "備考",
])
for rec in recipes:
    ws_rcp.append([rec[0], rec[1], rec[2], rec[3], None, None, None, rec[4]])

# ヘッダー
for c in range(1, 9):
    cell = ws_rcp.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# 各行に式を入れる
for i, rec in enumerate(recipes):
    r = i + 2
    # E: レシピ材料マスタからVLOOKUPで税抜単位単価
    ws_rcp.cell(row=r, column=5, value=f'=IFERROR(VLOOKUP(B{r},{SUB_RANGE_PRICE},6,FALSE),0)')
    # F: 材料費 = 分量 × 税抜単位単価
    ws_rcp.cell(row=r, column=6, value=f'=IFERROR(C{r}*E{r},0)')
    # G: 出来上がり寄与量（cc/ml/g のみ加算対象。個・本・缶・株・粒・pc は 0）
    ws_rcp.cell(row=r, column=7, value=f'=IF(OR(D{r}="cc",D{r}="ml",D{r}="g"),IFERROR(C{r},0),0)')
    ws_rcp.cell(row=r, column=3).number_format = '#,##0.00;(#,##0.00);-'
    ws_rcp.cell(row=r, column=5).number_format = '#,##0.0000;(#,##0.0000);-'
    ws_rcp.cell(row=r, column=6).number_format = '#,##0.00;(#,##0.00);-'
    ws_rcp.cell(row=r, column=7).number_format = '#,##0.00;(#,##0.00);-'

# 粒マスタード行: 分量を「ハンバーグソースの他材料合計×5%」にする
# 該当行を特定
for i, rec in enumerate(recipes):
    if rec[0] == "ハンバーグソース" and rec[1] == "粒マスタード":
        mr = i + 2
        # 他のハンバーグソース行の出来上がり寄与量(G列)合計 × 0.05
        # SUMIFS(G:G, A:A, "ハンバーグソース", B:B, "<>粒マスタード") の代わり
        ws_rcp.cell(row=mr, column=3, value='=SUMIFS(G:G,A:A,"ハンバーグソース",B:B,"<>粒マスタード")*0.05')
        break

ws_rcp.column_dimensions["A"].width = 22
ws_rcp.column_dimensions["B"].width = 22
ws_rcp.column_dimensions["C"].width = 10
ws_rcp.column_dimensions["D"].width = 7
ws_rcp.column_dimensions["E"].width = 20
ws_rcp.column_dimensions["F"].width = 14
ws_rcp.column_dimensions["G"].width = 18
ws_rcp.column_dimensions["H"].width = 30

RECIPE_LAST_ROW = 1 + len(recipes)
RECIPE_A_RANGE = f"ソースレシピマスタ!$A$2:$A${RECIPE_LAST_ROW}"
RECIPE_F_RANGE = f"ソースレシピマスタ!$F$2:$F${RECIPE_LAST_ROW}"
RECIPE_G_RANGE = f"ソースレシピマスタ!$G$2:$G${RECIPE_LAST_ROW}"

# -------------- ソース定義マスタ --------------
ws_sdef = wb.create_sheet("ソース定義マスタ")
# 各ソースの: バッチ材料費合計 / バッチ出来上がり量 / 単位原価(税抜・円/cc or 円/g) / 1食使用量基準
sauce_defs = [
    # name, batch yield override(cc), 基準（"batch" or "per_serving"）, 備考
    ("ハンバーグソース", None, "batch", "バッチ合計（味醂+醤油+水+粒マスタード等）の合算をccベースで使用"),
    ("ブールブランソース", 30, "per_serving", "1食=30cc 完成。レシピは1食分で記載"),
    ("オニオンソース", None, "batch", "オニオンステーキソース2L + ニンニク醤油ソース300cc"),
    ("ピクルス", 9200, "batch", "出来上がり9,200g＝生11,600g→可食部10,670g(廃棄率反映)→塩漬け×85%→水晒し×105%→水切り×97%。調味液は最終的に切るため重量加算なし"),
    ("ポリネシアンソース", None, "batch", "サラダ油+ケチャップ+白ワイン+醤油の合算をccベースで使用"),
    ("ソースジャンジャンブル", None, "batch", "液体材料の合計をccベースで使用"),
    ("エバラおろしのたれ", None, "market", "市販品。1Lあたりの単位原価で計算（業者単価で自動更新）"),
    ("粒マスタード単品", None, "condiment", "そのまま提供。g単位"),
    # サブレシピ（サイド用）
    ("サラダベース", None, "batch", "出来上がり量は C列に実測値を入力（個・束だけのため自動合計不可）"),
    ("ポテトサラダ", None, "batch", "レシピ未受領。出来上がり量と材料を後で入力"),
    ("蒸し野菜", 10280, "batch", "蒸し後重量(g)＝生重量12,850g×80%歩留まり仮算。キャベツ2×1200+白菜0.5×2500+にんじん3×180+玉ねぎ3×220+もやし8000"),
]
ws_sdef.append([
    "ソース名", "バッチ材料費（税抜・円）",
    "バッチ出来上がり量（cc/g 上書き可）",
    "出来上がり量（最終）", "単位原価（税抜・円/cc）",
    "基準", "備考",
])
for s in sauce_defs:
    ws_sdef.append([s[0], None, s[1], None, None, s[2], s[3]])

for c in range(1, 8):
    cell = ws_sdef.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for i, s in enumerate(sauce_defs):
    r = i + 2
    name = s[0]
    # B: バッチ材料費 = SUMIF(レシピマスタ A=ソース名, F=材料費)
    ws_sdef.cell(row=r, column=2,
        value=f'=SUMIFS({RECIPE_F_RANGE},{RECIPE_A_RANGE},A{r})')
    # D: 出来上がり量（最終）= 上書き(C)があればそれ、なければレシピのSUMIF
    ws_sdef.cell(row=r, column=4,
        value=f'=IF(ISNUMBER(C{r}),C{r},SUMIFS({RECIPE_G_RANGE},{RECIPE_A_RANGE},A{r}))')
    # E: 単位原価 = B / D
    ws_sdef.cell(row=r, column=5, value=f'=IFERROR(B{r}/D{r},0)')
    # スタイル
    ws_sdef.cell(row=r, column=2).number_format = '#,##0.00;(#,##0.00);-'
    ws_sdef.cell(row=r, column=3).fill = INPUT_FILL
    ws_sdef.cell(row=r, column=3).number_format = '#,##0.00;(#,##0.00);-'
    ws_sdef.cell(row=r, column=4).number_format = '#,##0.00;(#,##0.00);-'
    ws_sdef.cell(row=r, column=5).number_format = '#,##0.0000;(#,##0.0000);-'

ws_sdef.column_dimensions["A"].width = 22
ws_sdef.column_dimensions["B"].width = 22
ws_sdef.column_dimensions["C"].width = 26
ws_sdef.column_dimensions["D"].width = 18
ws_sdef.column_dimensions["E"].width = 22
ws_sdef.column_dimensions["F"].width = 12
ws_sdef.column_dimensions["G"].width = 40

SDEF_LAST_ROW = 1 + len(sauce_defs)
SAUCE_DEF_RANGE = f"ソース定義マスタ!$A$2:$E${SDEF_LAST_ROW}"

# サイドマスタの行参照（VLOOKUP用）— サイドマスタ範囲: A2:C8
SIDE_RANGE = "サイドマスタ!$A$2:$C$8"
ING_RANGE = "食材マスタ!$A$2:$F$17"
DRINK_RANGE = "飲料マスタ!$A$2:$D$5"
MAT_RANGE = "資材マスタ!$A$2:$C$10"

# -------------- メニューマスタ --------------
ws_menu = wb.create_sheet("メニューマスタ")

# Menu rows: (id, name, category, price, ing1, qty1, ing2, qty2, drink_id, side flags)
# side flags: (tsukemono, potato, tamago, rice, kabocha, ko_ebi, oo_ebi)
M = []
def m(*args): M.append(args)
# 黒弁当 (1-12): 漬物・ポテト・卵・かぼちゃ 全付与。ご飯=ID 9 以外。小エビ=ID 7,8 以外。大エビ頭=ID 7,8 のみ。
m(1, "赤城牛のハンバーグ弁当","黒弁当",2000,"150gバーグ",1,"",0,"none",1,1,1,1,1,1,0)
m(2, "特選ハラミのステーキ弁当","黒弁当",2160,"ハラミ",105,"",0,"none",1,1,1,1,1,1,0)
m(3, "ローストビーフ弁当","黒弁当",2160,"ロービー",85,"",0,"none",1,1,1,1,1,1,0)
m(4, "特選ハラミステーキとハンバーグの弁当","黒弁当",2200,"ハラミ",52,"150gバーグ",0.5,"none",1,1,1,1,1,1,0)
m(5, "特選ハラミステーキとローストビーフの弁当","黒弁当",2300,"ハラミ",52,"ロービー",42,"none",1,1,1,1,1,1,0)
m(6, "特選ハラミのステーキとイベリコ豚の弁当","黒弁当",2500,"ハラミ",52,"イベリコ",50,"pet_free",1,1,1,1,1,1,0)
m(7, "愛知県産真鯛の白ワイン蒸し弁当","黒弁当",2160,"鯛",1,"",0,"none",1,1,1,1,1,0,1)
m(8, "アトランティックサーモンのロースト弁当","黒弁当",2000,"サーモン",1,"",0,"none",1,1,1,1,1,0,1)
m(9, "低糖質弁当","黒弁当",2300,"大豆バーグ",1,"豆腐",0.33,"none",1,1,1,0,1,1,0)
m(10,"赤城牛ミスジとハラミのステーキ弁当","黒弁当",2800,"ミスジ",50,"ハラミ",52,"none",1,1,1,1,1,1,0)
m(11,"赤城牛ミスジのステーキ弁当","黒弁当",3000,"ミスジ",100,"",0,"pet_free",1,1,1,1,1,1,0)
m(12,"厳選和牛のステーキ弁当","黒弁当",3240,"厳選和牛",110,"",0,"pet_free",1,1,1,1,1,1,0)
# 新弁当 (13-25): 漬物・ポテト・卵・ご飯 全付与。かぼちゃ・小エビ=なし。大エビ頭=19,20,24。
m(13,"特選ハラミのステーキとローストビーフ弁当","新弁当",2100,"ハラミ",52,"ロービー",42,"none",1,1,1,1,0,0,0)
m(14,"特選ハラミのステーキ弁当","新弁当",1900,"ハラミ",105,"",0,"none",1,1,1,1,0,0,0)
m(15,"ローストビーフ弁当","新弁当",1900,"ロービー",85,"",0,"none",1,1,1,1,0,0,0)
m(16,"特選ハラミのステーキとハンバーグ弁当","新弁当",1800,"ハラミ",52,"120gバーグ",0.5,"none",1,1,1,1,0,0,0)
m(17,"赤城牛のハンバーグ弁当","新弁当",1700,"150gバーグ",1,"",0,"none",1,1,1,1,0,0,0)
m(18,"特製ハンバーグ弁当","新弁当",1500,"120gバーグ",1,"",0,"none",1,1,1,1,0,0,0)
m(19,"愛知県産真鯛の白ワイン蒸し弁当","新弁当",1900,"鯛",1,"",0,"none",1,1,1,1,0,0,1)
m(20,"アトランティックサーモンのロースト弁当","新弁当",1800,"サーモン",1,"",0,"none",1,1,1,1,0,0,1)
m(21,"特選豚肉と特製ハンバーグの弁当","新弁当",1620,"豚ヒレ",52,"120gバーグ",0.5,"none",1,1,1,1,0,0,0)
m(22,"特選豚ヒレ肉の弁当","新弁当",1500,"豚ヒレ",105,"",0,"none",1,1,1,1,0,0,0)
m(23,"特選豚ロースの弁当","新弁当",1500,"豚ロース",110,"",0,"none",1,1,1,1,0,0,0)
m(24,"カジキマグロと赤魚のソテー弁当","新弁当",1500,"カジキ",40,"赤魚",40,"none",1,1,1,1,0,0,1)
m(25,"赤城鶏のロースト弁当","新弁当",1400,"とりむね",120,"",0,"none",1,1,1,1,0,0,0)

headers = [
    "ID","メニュー名","分類","販売価格（税込）",
    "食材1","量1","食材2","量2",
    "ピクルス","ポテトサラダ","卵","ご飯","かぼちゃ","小エビ","有頭エビ",
    "飲料ID",
    "ソース1","ソース量1","ソース2","ソース量2",
    "主食材原価","サイド原価","資材原価","ソース原価","飲料原価",
    "合計原価（税抜）","販売価格（税抜）","原価率","利益率"
]
ws_menu.append(headers)
for c in range(1, len(headers)+1):
    cell = ws_menu.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# write menu rows + formulas
for i, row in enumerate(M):
    r = i + 2  # excel row
    (mid, name, cat, price, ing1, qty1, ing2, qty2, drink,
     f_tsuke, f_pote, f_tama, f_rice, f_kabo, f_koebi, f_ooebi) = row
    ws_menu.cell(row=r, column=1, value=mid)
    ws_menu.cell(row=r, column=2, value=name)
    ws_menu.cell(row=r, column=3, value=cat)
    ws_menu.cell(row=r, column=4, value=price)
    ws_menu.cell(row=r, column=5, value=ing1)
    ws_menu.cell(row=r, column=6, value=qty1)
    ws_menu.cell(row=r, column=7, value=ing2 if ing2 else None)
    ws_menu.cell(row=r, column=8, value=qty2 if ing2 else None)
    ws_menu.cell(row=r, column=9,  value=f_tsuke)
    ws_menu.cell(row=r, column=10, value=f_pote)
    ws_menu.cell(row=r, column=11, value=f_tama)
    ws_menu.cell(row=r, column=12, value=f_rice)
    ws_menu.cell(row=r, column=13, value=f_kabo)
    ws_menu.cell(row=r, column=14, value=f_koebi)
    ws_menu.cell(row=r, column=15, value=f_ooebi)
    ws_menu.cell(row=r, column=16, value=drink)

    # ソース入力欄(17-20)は空のままにしておく（ユーザーがメニューごとに指定）
    # Q=ソース1, R=ソース量1, S=ソース2, T=ソース量2

    # 主食材原価 (col 21 = U)
    f_ing = (
        f'=IFERROR(VLOOKUP(E{r},{ING_RANGE},6,FALSE)*F{r},0)'
        f'+IFERROR(VLOOKUP(G{r},{ING_RANGE},6,FALSE)*H{r},0)'
    )
    ws_menu.cell(row=r, column=21, value=f_ing)

    # サイド原価 (col 22 = V): 重量ベース
    # フラグ管理7サイド: flag × IF(分類=黒, 黒量, 新量) × 単位原価
    # 全付与4サイド: IF(分類=黒, 黒量, 新量) × 単位原価
    flag_sides = [
        ("ピクルス",     "I"),
        ("ポテトサラダ", "J"),
        ("卵",          "K"),
        ("ご飯",        "L"),
        ("かぼちゃ",     "M"),
        ("小エビ",      "N"),
        ("有頭エビ",     "O"),
    ]
    always_sides = ["サラダベース", "ミニトマト", "蒸し野菜", "ブロッコリー"]
    parts = []
    for nm, fc in flag_sides:
        # qty: IF(分類="黒弁当", VLOOKUP(name, side, 3), VLOOKUP(name, side, 4))
        # unit_price: VLOOKUP(name, side, 5)
        parts.append(
            f'IFERROR({fc}{r}*'
            f'IF($C{r}="黒弁当",IFERROR(VLOOKUP("{nm}",サイドマスタ!$A:$F,3,FALSE),0),IFERROR(VLOOKUP("{nm}",サイドマスタ!$A:$F,4,FALSE),0))'
            f'*IFERROR(VLOOKUP("{nm}",サイドマスタ!$A:$F,5,FALSE),0),0)'
        )
    for nm in always_sides:
        parts.append(
            f'IFERROR(IF($C{r}="黒弁当",IFERROR(VLOOKUP("{nm}",サイドマスタ!$A:$F,3,FALSE),0),IFERROR(VLOOKUP("{nm}",サイドマスタ!$A:$F,4,FALSE),0))'
            f'*IFERROR(VLOOKUP("{nm}",サイドマスタ!$A:$F,5,FALSE),0),0)'
        )
    ws_menu.cell(row=r, column=22, value="=" + "+".join(parts))

    # 資材原価 (col 23 = W)
    common_mat = (
        f'IFERROR(VLOOKUP("箸",{MAT_RANGE},3,FALSE),0)'
        f'+IFERROR(VLOOKUP("ワサガード",{MAT_RANGE},3,FALSE),0)'
        f'+IFERROR(VLOOKUP("帯",{MAT_RANGE},3,FALSE),0)'
        f'+IFERROR(VLOOKUP("シーザードレッシング",{MAT_RANGE},3,FALSE),0)'
        f'+IFERROR(VLOOKUP("おしぼり",{MAT_RANGE},3,FALSE),0)'
    )
    size_mat = (
        f'IF(C{r}="黒弁当",'
        f'IFERROR(VLOOKUP("容器（黒）",{MAT_RANGE},3,FALSE),0)+IFERROR(VLOOKUP("蓋（黒）",{MAT_RANGE},3,FALSE),0),'
        f'IFERROR(VLOOKUP("容器（新）",{MAT_RANGE},3,FALSE),0)+IFERROR(VLOOKUP("蓋（新）",{MAT_RANGE},3,FALSE),0))'
    )
    ws_menu.cell(row=r, column=23, value=f'={size_mat}+{common_mat}')

    # ソース原価 (col 24 = X): ソース定義マスタの単位原価(税抜) × 使用量(cc)
    ws_menu.cell(row=r, column=24, value=(
        f'=IFERROR(VLOOKUP(Q{r},{SAUCE_DEF_RANGE},5,FALSE)*R{r},0)'
        f'+IFERROR(VLOOKUP(S{r},{SAUCE_DEF_RANGE},5,FALSE)*T{r},0)'
    ))

    # 飲料原価 (col 25 = Y)
    ws_menu.cell(row=r, column=25, value=f'=IFERROR(VLOOKUP(P{r},{DRINK_RANGE},4,FALSE),0)')
    # 合計原価（税抜） (col 26 = Z) = 主食材 + サイド + 資材 + ソース + 飲料
    ws_menu.cell(row=r, column=26, value=f'=U{r}+V{r}+W{r}+X{r}+Y{r}')
    # 販売価格（税抜） (col 27 = AA)
    ws_menu.cell(row=r, column=27, value=f'=D{r}/1.08')
    # 原価率 (col 28 = AB)
    ws_menu.cell(row=r, column=28, value=f'=IFERROR(Z{r}/AA{r},0)')
    # 利益率 (col 29 = AC)
    ws_menu.cell(row=r, column=29, value=f'=IFERROR(1-AB{r},0)')

# メニューごとのソース割当（ヒアリングで埋めていく）
# 形式: ID -> [(ソース名1, 量1), (ソース名2, 量2)]  ※2本目なしなら省略
menu_sauces = {
    1: [("ハンバーグソース", 40)],
    2: [("オニオンソース", 30)],
    3: [("エバラおろしのたれ", 30)],
    4: [("オニオンソース", 20), ("ハンバーグソース", 30)],
    5: [("オニオンソース", 20), ("エバラおろしのたれ", 20)],
    6: [("オニオンソース", 20), ("粒マスタード単品", 15)],
    7: [("ブールブランソース", 30)],
    8: [("ブールブランソース", 30)],
    9: [("エバラおろしのたれ", 30)],
    10: [("エバラおろしのたれ", 20), ("オニオンソース", 20)],
    11: [("エバラおろしのたれ", 30)],
    12: [("エバラおろしのたれ", 30)],
    13: [("オニオンソース", 20), ("エバラおろしのたれ", 20)],
    14: [("オニオンソース", 30)],
    15: [("エバラおろしのたれ", 30)],
    16: [("オニオンソース", 20), ("ハンバーグソース", 25)],
    17: [("ハンバーグソース", 40)],
    18: [("ハンバーグソース", 35)],
    19: [("ブールブランソース", 30)],
    20: [("ブールブランソース", 30)],
    21: [("ポリネシアンソース", 20), ("ハンバーグソース", 25)],
    22: [("ポリネシアンソース", 30)],
    23: [("ソースジャンジャンブル", 30)],
    24: [("エバラおろしのたれ", 30)],
    25: [("ポリネシアンソース", 30)],
}
for i, row in enumerate(M):
    r = i + 2
    mid = row[0]
    sauces_for_menu = menu_sauces.get(mid, [])
    if len(sauces_for_menu) >= 1:
        ws_menu.cell(row=r, column=17, value=sauces_for_menu[0][0])  # Q ソース1
        ws_menu.cell(row=r, column=18, value=sauces_for_menu[0][1])  # R 量1
    if len(sauces_for_menu) >= 2:
        ws_menu.cell(row=r, column=19, value=sauces_for_menu[1][0])  # S ソース2
        ws_menu.cell(row=r, column=20, value=sauces_for_menu[1][1])  # T 量2

# number formats & ソース入力欄の薄黄色
for r in range(2, 2 + len(M)):
    # ソース入力欄 (Q,R,S,T) を薄黄色
    for c in [17, 18, 19, 20]:
        ws_menu.cell(row=r, column=c).fill = INPUT_FILL
    # 通貨 (販売価格・各原価)
    for c in [4, 21, 22, 23, 24, 25, 26, 27]:
        ws_menu.cell(row=r, column=c).number_format = '#,##0;(#,##0);-'
    # パーセント
    for c in [28, 29]:
        ws_menu.cell(row=r, column=c).number_format = '0.0%'

# column widths
widths = {1:5,2:34,3:8,4:12,5:14,6:7,7:14,8:7,
          9:6,10:6,11:6,12:6,13:7,14:7,15:8,16:10,
          17:14,18:8,19:14,20:8,
          21:12,22:12,23:12,24:12,25:12,26:14,27:14,28:9,29:9}
for col, w in widths.items():
    ws_menu.column_dimensions[get_column_letter(col)].width = w
ws_menu.freeze_panes = "E2"

# 行高さを少しゆったり
ws_menu.row_dimensions[1].height = 32
# 合計原価列を金色背景で強調
for r in range(2, 2 + len(M)):
    ws_menu.cell(row=r, column=26).fill = TOTAL_FILL
    ws_menu.cell(row=r, column=26).font = Font(name=FONT, bold=True)
# 罫線を全データセルに
for r in range(1, 2 + len(M)):
    for c in range(1, 30):
        ws_menu.cell(row=r, column=c).border = BORDER
# 原価率列に条件付き書式
menu_last = 1 + len(M)
ws_menu.conditional_formatting.add(f"AB2:AB{menu_last}", RULE_DANGER)
ws_menu.conditional_formatting.add(f"AB2:AB{menu_last}", RULE_WARN)
ws_menu.conditional_formatting.add(f"AB2:AB{menu_last}", RULE_GREEN)
# 利益率列にも色付け（緑系で）
ws_menu.conditional_formatting.add(f"AC2:AC{menu_last}",
    ColorScaleRule(start_type='num', start_value=0.4, start_color="FFC7CE",
                   mid_type='num', mid_value=0.6, mid_color="FFEB9C",
                   end_type='num', end_value=0.8, end_color="C6EFCE"))

# -------------- サマリー --------------
ws_sum = wb.create_sheet("サマリー")
# タイトルバナー
ws_sum.merge_cells("A1:D1")
ws_sum["A1"] = "Bistro Knocks 原価サマリー"
ws_sum["A1"].font = TITLE_FONT
ws_sum["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws_sum.row_dimensions[1].height = 38
ws_sum["A1"].border = THICK_BOTTOM
ws_sum["B1"].border = THICK_BOTTOM
ws_sum["C1"].border = THICK_BOTTOM
ws_sum["D1"].border = THICK_BOTTOM

# 3列構造で並べる: 項目 | 黒弁当 | 新弁当 | 全メニュー
header_row = 3
ws_sum.cell(row=header_row, column=1, value="指標")
ws_sum.cell(row=header_row, column=2, value="黒弁当")
ws_sum.cell(row=header_row, column=3, value="新弁当")
ws_sum.cell(row=header_row, column=4, value="全メニュー")
for c in range(1, 5):
    cell = ws_sum.cell(row=header_row, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER

# 各指標行
rows_data = [
    ("平均販売価格（税込）", "currency",
        '=AVERAGEIF(メニューマスタ!C2:C26,"黒弁当",メニューマスタ!D2:D26)',
        '=AVERAGEIF(メニューマスタ!C2:C26,"新弁当",メニューマスタ!D2:D26)',
        '=AVERAGE(メニューマスタ!D2:D26)'),
    ("平均販売価格（税抜）", "currency",
        '=AVERAGEIFS(メニューマスタ!AA2:AA26,メニューマスタ!C2:C26,"黒弁当")',
        '=AVERAGEIFS(メニューマスタ!AA2:AA26,メニューマスタ!C2:C26,"新弁当")',
        '=AVERAGE(メニューマスタ!AA2:AA26)'),
    ("平均原価（税抜）", "currency",
        '=AVERAGEIF(メニューマスタ!C2:C26,"黒弁当",メニューマスタ!Z2:Z26)',
        '=AVERAGEIF(メニューマスタ!C2:C26,"新弁当",メニューマスタ!Z2:Z26)',
        '=AVERAGE(メニューマスタ!Z2:Z26)'),
    ("平均原価率", "pct",
        '=AVERAGEIF(メニューマスタ!C2:C26,"黒弁当",メニューマスタ!AB2:AB26)',
        '=AVERAGEIF(メニューマスタ!C2:C26,"新弁当",メニューマスタ!AB2:AB26)',
        '=AVERAGE(メニューマスタ!AB2:AB26)'),
    ("平均利益率", "pct",
        '=AVERAGEIF(メニューマスタ!C2:C26,"黒弁当",メニューマスタ!AC2:AC26)',
        '=AVERAGEIF(メニューマスタ!C2:C26,"新弁当",メニューマスタ!AC2:AC26)',
        '=AVERAGE(メニューマスタ!AC2:AC26)'),
]
for i, (label, fmt, vb, vn, va) in enumerate(rows_data):
    r = header_row + 1 + i
    ws_sum.cell(row=r, column=1, value=label).font = BASE_FONT
    ws_sum.cell(row=r, column=2, value=vb)
    ws_sum.cell(row=r, column=3, value=vn)
    ws_sum.cell(row=r, column=4, value=va)
    for c in range(2, 5):
        cell = ws_sum.cell(row=r, column=c)
        cell.alignment = Alignment(horizontal="right")
        if fmt == "currency":
            cell.number_format = '¥#,##0;(¥#,##0);-'
        else:
            cell.number_format = '0.0%'
    # 縞模様
    if i % 2 == 1:
        for c in range(1, 5):
            ws_sum.cell(row=r, column=c).fill = SECTION_FILL
    for c in range(1, 5):
        ws_sum.cell(row=r, column=c).border = BORDER

# 原価率行（行5）に条件付き書式
ws_sum.conditional_formatting.add(f"B{header_row+4}:D{header_row+4}", RULE_DANGER)
ws_sum.conditional_formatting.add(f"B{header_row+4}:D{header_row+4}", RULE_WARN)
ws_sum.conditional_formatting.add(f"B{header_row+4}:D{header_row+4}", RULE_GREEN)

ws_sum.column_dimensions["A"].width = 32
ws_sum.column_dimensions["B"].width = 16
ws_sum.column_dimensions["C"].width = 16
ws_sum.column_dimensions["D"].width = 16

# -------------- README --------------
ws_doc = wb.create_sheet("使い方", 0)
# タイトルバナー
ws_doc.merge_cells("A1:C1")
ws_doc["A1"] = "Bistro Knocks  原価管理シート"
ws_doc["A1"].font = TITLE_FONT
ws_doc["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws_doc.row_dimensions[1].height = 42
# サブタイトル
ws_doc.merge_cells("A2:C2")
ws_doc["A2"] = "Cost Master — 業者単価を入れるだけで全メニューの原価率／利益率が自動算出"
ws_doc["A2"].font = SUBTITLE_FONT
ws_doc["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws_doc.row_dimensions[2].height = 22
# 区切り線
for col in ["A","B","C"]:
    ws_doc[f"{col}2"].border = THICK_BOTTOM

# 凡例（行3-5）
ws_doc["A4"] = "凡例"
ws_doc["A4"].font = SUBTITLE_FONT
ws_doc["B5"] = "  入力欄  "
ws_doc["B5"].fill = INPUT_FILL
ws_doc["B5"].border = BORDER
ws_doc["B5"].alignment = Alignment(horizontal="center")
ws_doc["C5"] = "業者単価などを手入力（薄アイボリー）"
ws_doc["C5"].font = BASE_FONT
ws_doc["B6"] = "  自動計算  "
ws_doc["B6"].fill = PatternFill("solid", start_color=COLOR_COMPUTED_BG)
ws_doc["B6"].border = BORDER
ws_doc["B6"].alignment = Alignment(horizontal="center")
ws_doc["C6"] = "関数で算出（白）"
ws_doc["C6"].font = BASE_FONT
ws_doc["B7"] = "  合計／重要  "
ws_doc["B7"].fill = TOTAL_FILL
ws_doc["B7"].border = BORDER
ws_doc["B7"].alignment = Alignment(horizontal="center")
ws_doc["C7"] = "合計原価などの重要セル（淡ゴールド）"
ws_doc["C7"].font = BASE_FONT

ws_doc["A1"] = "Bistro Knocks  原価管理シート"  # ensure value persists
ws_doc["A1"].font = TITLE_FONT
ws_doc["A1"].alignment = Alignment(horizontal="left", vertical="center")
notes = [
    "",
    "【税の前提】",
    "・販売価格は税込で運用、業者からの単価は税抜で入る前提です。",
    "・原価率・利益率は飲食業の慣行に合わせ「税抜ベース」で算出します。",
    "・販売価格（税抜）= 販売価格（税込）÷ 1.08（軽減税率 / テイクアウト弁当）。",
    "・各マスタの単価入力欄は全て『税抜』で入力してください。",
    "・食材の使用量は『掃除前重量』（業者単価と同じベース）で記載。歩留まりは別途反映しません。",
    "",
    "【使い方】",
    "1. 業者から単価が届いたら、以下の薄黄色セルに入力してください：",
    "   ・食材マスタ シート C列＝仕入単価（税込）、D列＝仕入単位量（例: 100g単位なら 100）",
    "     →E列「税込/単位」と F列「税抜/単位」が自動計算され、原価計算ではF列が使われます。",
    "   ・飲料マスタ シート D列（pet_free / pet / pack の3件）",
    "   ・サイドマスタ シート C列（7品目: ピクルス・ポテト・卵・ご飯・かぼちゃ・小エビ・有頭エビ）",
    "   ・資材マスタ シート C列（9品目: 容器・蓋（黒/新別）・箸・ワサガード・帯・シーザードレッシング・おしぼり）",
    "   ・レシピ材料マスタ シート C〜D列（ソース材料の仕入単価（税込）・仕入単位量）",
    "   ・ソースレシピマスタ シート（レシピ既入力。材料変更時のみ編集）",
    "   ・ソース定義マスタ シート C列（任意・出来上がり量の上書き。空ならレシピ合計）",
    "   ・メニューマスタ シート Q列（ソース名）・R列（1食あたり使用量cc）／ハーフはS,T列に2本目",
    "2. 入力すると メニューマスタ の「合計原価／原価率／利益率」が自動計算されます。",
    "3. サマリー シートに分類別の平均値が出ます。",
    "",
    "【シート構成】",
    "・食材マスタ：主食材16品目の単価",
    "・飲料マスタ：飲料4種（noneは0固定）",
    "・サイドマスタ：1食あたりの固定原価（サイド7品）",
    "・資材マスタ：包装・容器など資材9品（黒弁当/新弁当でサイズ別あり）",
    "・レシピ材料マスタ：ソース・ピクルスのサブ材料（味醂・醤油・きのこ等）の業者単価",
    "・ソースレシピマスタ：各ソースのバッチレシピ（1行=ソース×材料）。材料費・出来上がり量を自動計算",
    "・ソース定義マスタ：ソースごとのバッチ材料費合計・出来上がり量・単位原価（円/cc）",
    "・メニューマスタ：25メニューの原価計算（関数） / Q〜T列にメニュー別ソースを指定",
    "・サマリー：分類別の平均販売価格・原価・原価率・利益率",
    "",
    "【サイド付与ルール】",
    "・黒弁当(ID 1-12): ピクルス・ポテト・卵・かぼちゃ 全付与。ご飯はID 9(低糖質)以外。小エビは魚系(ID 7,8)以外。有頭エビは鯛(7)・サーモン(8)のみ。",
    "・新弁当(ID 13-25): ピクルス・ポテト・卵・ご飯 全付与。有頭エビは鯛(19)・サーモン(20)・カジキ赤魚(24)のみ。",
    "・付与有無は メニューマスタ の I〜O列 (1=付与, 0=なし) で個別に調整できます。",
    "",
    "薄黄色＝業者単価待ちの入力セル",
]
# notes は凡例ブロック(行3-7)の下に配置
NOTE_START = 9
for i, line in enumerate(notes, start=NOTE_START):
    cell = ws_doc.cell(row=i, column=1, value=line)
    # セクション見出し（【...】）はサブタイトルスタイル
    if line.startswith("【") and line.endswith("】"):
        ws_doc.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        cell.font = SUBTITLE_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_doc.row_dimensions[i].height = 22
    elif line.strip() == "":
        pass
    else:
        ws_doc.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        cell.font = BASE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

ws_doc.column_dimensions["A"].width = 28
ws_doc.column_dimensions["B"].width = 16
ws_doc.column_dimensions["C"].width = 80

# シートのタブ色
ws_doc.sheet_properties.tabColor = COLOR_HEADER_BG   # 使い方=チャコール
ws_ing.sheet_properties.tabColor = "A9CCE3"          # 食材=スカイブルー
ws_drink.sheet_properties.tabColor = "A9CCE3"
ws_side.sheet_properties.tabColor = "A9CCE3"
ws_mat.sheet_properties.tabColor = "A9CCE3"
ws_sub.sheet_properties.tabColor = "A9CCE3"          # レシピ材料=スカイブルー
ws_rcp.sheet_properties.tabColor = "D5C5A4"          # レシピ計算系=ウォームグレー
ws_sdef.sheet_properties.tabColor = "D5C5A4"
ws_menu.sheet_properties.tabColor = COLOR_ACCENT     # メニュー=ゴールド（主要出力）
ws_sum.sheet_properties.tabColor = COLOR_ACCENT      # サマリー=ゴールド（主要出力）

# 1ページ目に表示するシート = 使い方
wb.active = wb.sheetnames.index("使い方")

wb.save(OUT)
print("Saved:", OUT)
