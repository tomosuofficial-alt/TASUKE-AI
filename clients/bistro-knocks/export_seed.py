#!/usr/bin/env python3
"""Export Bistro Knocks cost master data as stack-agnostic seed.

Reads bistro_knocks_cost_master.xlsx and outputs:
- seed/*.json  : raw entity data (drop into any web app)
- types.ts     : TypeScript interfaces describing the schema
- business-logic.md : cost calculation spec

Usage:
    python3 export_seed.py
"""
from openpyxl import load_workbook
import json
from pathlib import Path

BASE = Path("/Users/takashiouchi/Developer/TASUKE-AI/.claude/worktrees/peaceful-tu-b43503/clients/bistro-knocks")
WB_PATH = BASE / "bistro_knocks_cost_master.xlsx"
OUT_DIR = BASE / "seed"
OUT_DIR.mkdir(exist_ok=True)


def cell_or_none(v):
    """Convert openpyxl value to clean primitive; formula cells become None."""
    if v is None:
        return None
    if isinstance(v, str) and v.startswith("="):
        return None  # formula → computed by app
    return v


def export_food_ingredients(wb):
    ws = wb['食材マスタ']
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        rows.append({
            "name": name,
            "unit": ws.cell(r, 2).value,
            "purchase_price_incl_tax": cell_or_none(ws.cell(r, 3).value),
            "purchase_unit_qty": cell_or_none(ws.cell(r, 4).value),
            "note": ws.cell(r, 7).value,
        })
    return rows


def export_drinks(wb):
    ws = wb['飲料マスタ']
    rows = []
    for r in range(2, ws.max_row + 1):
        drink_id = ws.cell(r, 1).value
        if not drink_id:
            continue
        rows.append({
            "id": drink_id,
            "name": ws.cell(r, 2).value,
            "sales_price_incl_tax": ws.cell(r, 3).value,
            "cost_excl_tax": cell_or_none(ws.cell(r, 4).value),
            "note": ws.cell(r, 5).value,
        })
    return rows


def export_sides(wb):
    ws = wb['サイドマスタ']
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        rows.append({
            "name": name,
            "unit": ws.cell(r, 2).value,
            "qty_black_bento": cell_or_none(ws.cell(r, 3).value),
            "qty_new_bento": cell_or_none(ws.cell(r, 4).value),
            "unit_price_excl_tax": cell_or_none(ws.cell(r, 5).value),
            "note": ws.cell(r, 6).value,
        })
    return rows


def export_packaging(wb):
    ws = wb['資材マスタ']
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        rows.append({
            "name": name,
            "category": ws.cell(r, 2).value,
            "price_excl_tax": cell_or_none(ws.cell(r, 3).value),
            "applies_to": ws.cell(r, 4).value,
        })
    return rows


def export_recipe_ingredients(wb):
    ws = wb['レシピ材料マスタ']
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        rows.append({
            "name": name,
            "unit": ws.cell(r, 2).value,
            "purchase_price_incl_tax": cell_or_none(ws.cell(r, 3).value),
            "purchase_unit_qty": cell_or_none(ws.cell(r, 4).value),
            "note": ws.cell(r, 7).value,
        })
    return rows


def export_sub_recipes(wb):
    """Sauces, pickle, salad-base, steamed-veg recipes."""
    ws = wb['ソースレシピマスタ']
    rows = []
    for r in range(2, ws.max_row + 1):
        recipe_name = ws.cell(r, 1).value
        ingredient = ws.cell(r, 2).value
        if not recipe_name or not ingredient:
            continue
        qty_val = ws.cell(r, 3).value
        # 粒マスタード行は formula なので None になる
        rows.append({
            "recipe_name": recipe_name,
            "ingredient_name": ingredient,
            "qty": cell_or_none(qty_val) if not (isinstance(qty_val, str) and qty_val.startswith("=")) else "DYNAMIC:5%_OF_TOTAL",
            "unit": ws.cell(r, 4).value,
            "note": ws.cell(r, 8).value,
        })
    return rows


def export_sub_recipe_definitions(wb):
    ws = wb['ソース定義マスタ']
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        rows.append({
            "name": name,
            "batch_yield_override": cell_or_none(ws.cell(r, 3).value),
            "basis": ws.cell(r, 6).value,
            "note": ws.cell(r, 7).value,
        })
    return rows


def export_menus(wb):
    ws = wb['メニューマスタ']
    rows = []
    for r in range(2, ws.max_row + 1):
        mid = ws.cell(r, 1).value
        if not mid:
            continue
        side_flags = {
            "pickles": ws.cell(r, 9).value,
            "potato_salad": ws.cell(r, 10).value,
            "egg": ws.cell(r, 11).value,
            "rice": ws.cell(r, 12).value,
            "pumpkin": ws.cell(r, 13).value,
            "small_shrimp": ws.cell(r, 14).value,
            "head_on_shrimp": ws.cell(r, 15).value,
        }
        sauces = []
        if ws.cell(r, 17).value:
            sauces.append({"name": ws.cell(r, 17).value, "qty_cc": ws.cell(r, 18).value})
        if ws.cell(r, 19).value:
            sauces.append({"name": ws.cell(r, 19).value, "qty_cc": ws.cell(r, 20).value})
        ingredients = []
        if ws.cell(r, 5).value:
            ingredients.append({"name": ws.cell(r, 5).value, "qty": ws.cell(r, 6).value})
        if ws.cell(r, 7).value:
            ingredients.append({"name": ws.cell(r, 7).value, "qty": ws.cell(r, 8).value})
        rows.append({
            "id": mid,
            "name": ws.cell(r, 2).value,
            "category": ws.cell(r, 3).value,  # 黒弁当 / 新弁当
            "sales_price_incl_tax": ws.cell(r, 4).value,
            "ingredients": ingredients,
            "side_flags": side_flags,
            "drink_id": ws.cell(r, 16).value,
            "sauces": sauces,
        })
    return rows


def main():
    wb = load_workbook(WB_PATH, data_only=False)

    exports = {
        "01_food_ingredients.json": export_food_ingredients(wb),
        "02_drinks.json": export_drinks(wb),
        "03_sides.json": export_sides(wb),
        "04_packaging_materials.json": export_packaging(wb),
        "05_recipe_ingredients.json": export_recipe_ingredients(wb),
        "06_sub_recipes.json": export_sub_recipes(wb),
        "07_sub_recipe_definitions.json": export_sub_recipe_definitions(wb),
        "08_menus.json": export_menus(wb),
    }

    for filename, data in exports.items():
        path = OUT_DIR / filename
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2))
        print(f"  ✓ {path.relative_to(BASE)}  ({len(data)} entries)")

    # Generate combined index
    combined = {k.replace(".json", "").split("_", 1)[1]: v for k, v in exports.items()}
    (OUT_DIR / "all.json").write_text(json.dumps(combined, ensure_ascii=False, indent=2))
    print(f"  ✓ seed/all.json  (combined)")

    print(f"\nDone. {len(exports)} files written to {OUT_DIR.relative_to(BASE)}/")


if __name__ == "__main__":
    main()
